 
import openpyxl
import pandas as pd

def main():
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("./huy.xlsx")

    # Define variable to read sheet
    wb = dataframe.active
    
    for row in range(1, wb.max_row):
        if wb[row+1][8].value == None:
            continue
        s = wb[row+1][8].value.split('\n')

        wb[row+1][8].value = ''
        for i in range(0,len(s)):
            s[i] = s[i].strip()
            if len(s[i]) == 0:
                continue
            if s[i][0] == '+' and i != 0:
                wb[row+1][8].value += '\n'
            wb[row+1][8].value += s[i]

            if s[i][-1] == ':' or s[i][0] == '+':
                wb[row+1][8].value += ''
            else:
                if i != len(s)-1:
                    wb[row+1][8].value += '\n'

    dataframe.save(filename = "new.xlsx")

def main2():
    # Define variable to load the dataframe
    dataframe = openpyxl.load_workbook("./new.xlsx")

    # Define variable to read sheet
    wb = dataframe.active
    
    for row in range(1, wb.max_row):
        if wb[row+1][8].value == None:
            continue
        s = wb[row+1][8].value

        wb[row+1][8].value = s[0]
        for i in range(1,len(s)):
            if s[i-1].islower() and s[i].isupper():
                wb[row+1][8].value += ', '
            wb[row+1][8].value += s[i]

    dataframe.save(filename = "new2.xlsx")

def exe2txt():
    data = pd.read_excel("./data/excel/10_groupitem (2).xlsx")

    txt = 'I. Các loại sản phẩm'
    ex_txt = ''
    cnt = 0
    cnt_pro = 0
    for row in data.index:
        if row == 0 or data['GROUP_PRODUCT_NAME'][row] != data['GROUP_PRODUCT_NAME'][row-1]:
            if cnt_pro != 0:
                txt += '\n\tSố lượng sản phẩm là: ' + str(cnt_pro)
            txt += ex_txt

            cnt += 1
            txt += '\n' + str(cnt) + '. ' + data['GROUP_PRODUCT_NAME'][row]
            ex_txt = ''
            cnt_pro = 0
            
        ex_txt += '\n\t- ' + str(data['PRODUCT_INFO_ID'][row]) + ': ' + data['PRODUCT_NAME'][row]
        cnt_pro +=1
    txt += '\n\tSố lượng sản phẩm là: ' + str(cnt_pro)
    txt += ex_txt
    
    txt += '\nII. Thông tin sản phẩm'
    for row in data.index:
        txt += '\n' + str(row+1) + '. ' + data['GROUP_PRODUCT_NAME'][row] + ': ' + str(data['PRODUCT_INFO_ID'][row])
        txt += '\n\t' + '- Tên sản phẩm: ' + data['PRODUCT_NAME'][row]

        # if wb[row+1][4].value != None:
        #     txt += '\n' + '- Giá: ' + str(wb[row+1][4].value)
        # else:
        #     txt += '\n' + '- Giá: hiện tại giá thành chưa được cập nhật'
        txt += '\n\t' + '- Mô tả: ' + data['SHORT_DESCRIPTION'][row]
        
        # txt += '\n' + '- Thông số kỹ thuật:'
        # if wb[row+1][8].value != None:
        #     s = wb[row+1][8].value.split('\n')
        #     for i in s:
        #         txt += '\n\t' + i
        # else:
        #     txt += 'hiện tại thông số kỹ thuật chưa được cập nhật'
        
        # if wb[row+1][9].value != None:
        #     txt += '\n' + '- Giới thiệu sản phẩm:'
        #     s = wb[row+1][9].value.split('\n')
        #     for i in s:
        #         if len(i) == 0:
        #             continue
        #         txt += '\n\t' + i

    f = open("product_detail.txt", "w", encoding="utf-8")
    f.write(txt)
    f.close()

exe2txt()